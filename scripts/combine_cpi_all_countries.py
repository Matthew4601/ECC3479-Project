"""
Combine CPI data for Canada (Statistics Canada 18-10-0256-01),
UK (ONS CPIH index), and USA (BLS CPI components).

Sources:
  data/raw/1810025601-eng(CAN CPI).xlsx  — Canada core CPI measures, wide format, Jan 2020–Feb 2026
  data/raw/UK_CPI_index.csv              — UK CPIH all-items index (2015=100), monthly
  data/raw/USA CPI.xlsx                  — USA CPI 18-component series (YoY % change), monthly

Overlapping period (inner join): Jan 2020 – Feb 2026 (74 months, all three sources match exactly)
Output: data/clean/combined_cpi_all_countries.csv
"""

import re
from pathlib import Path
from datetime import datetime

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
CAN_PATH = ROOT / "data" / "raw" / "1810025601-eng(CAN CPI).xlsx"
UK_PATH  = ROOT / "data" / "raw" / "UK_CPI_index.csv"
USA_PATH = ROOT / "data" / "raw" / "USA CPI.xlsx"
OUT_PATH = ROOT / "data" / "clean" / "combined_cpi_all_countries.csv"

MONTH_MAP = {
    "January": 1, "February": 2, "March": 3, "April": 4,
    "May": 5, "June": 6, "July": 7, "August": 8,
    "September": 9, "October": 10, "November": 11, "December": 12,
}

UK_MONTH_MAP = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}

MONTHLY_RE = re.compile(r"^(\d{4}) ([A-Z]{3})$")

# Short slugs for each Canadian measure row (rows 12–26 in the sheet, 0-indexed)
CAN_MEASURE_SLUGS = [
    "can_cpi_common_yoy_pct",
    "can_cpi_median_yoy_pct",
    "can_cpi_trim_yoy_pct",
    "can_cpi_median_idx198901",
    "can_cpi_trim_idx198901",
    "can_cpi_ex8volatile_idx2002",
    "can_cpi_ex8volatile_notax_idx2002",
    "can_cpi_ex_indirect_tax_idx2002",
    "can_cpi_ex_food_energy_tax_idx2002",
    "can_cpi_ex8volatile_sa_notax_idx2002",
    "can_cpi_ex8volatile_sa_idx2002",
    "can_cpi_ex_indirect_tax_sa_idx2002",
    "can_cpi_ex_food_energy_tax_sa_idx2002",
]


def load_canada() -> pd.DataFrame:
    wb = openpyxl.load_workbook(CAN_PATH)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Row index 10 = header; cols 1+ are date strings like "January 2020"
    date_row = rows[10]
    date_cols = {}  # col_index -> datetime
    for i, val in enumerate(date_row[1:], start=1):
        if isinstance(val, str) and val in MONTH_MAP:
            pass  # shouldn't happen
        if isinstance(val, str):
            parts = val.split()
            if len(parts) == 2 and parts[0] in MONTH_MAP:
                date_cols[i] = datetime(int(parts[1]), MONTH_MAP[parts[0]], 1)

    # Data rows start at index 12; collect only the measure rows (stop before footnotes at row 31)
    data_rows = [r for r in rows[12:31] if r[0] is not None]

    records = {dt: {"date": dt} for dt in date_cols.values()}
    for measure_row, slug in zip(data_rows, CAN_MEASURE_SLUGS):
        for col_idx, dt in date_cols.items():
            val = measure_row[col_idx]
            records[dt][slug] = float(val) if val is not None else None

    return pd.DataFrame(list(records.values())).sort_values("date").reset_index(drop=True)


def load_uk() -> pd.DataFrame:
    rows = []
    with open(UK_PATH, encoding="utf-8") as f:
        for line in f:
            line = line.strip().strip('"')
            parts = line.split('","')
            if len(parts) != 2:
                continue
            label, value = parts[0].strip(), parts[1].strip()
            m = MONTHLY_RE.match(label)
            if m:
                year, mon = int(m.group(1)), UK_MONTH_MAP[m.group(2)]
                rows.append({"date": datetime(year, mon, 1), "uk_cpih_all_items_idx2015": float(value)})
    return pd.DataFrame(rows).sort_values("date").reset_index(drop=True)


def load_usa() -> pd.DataFrame:
    wb = openpyxl.load_workbook(USA_PATH)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    header = rows[0]
    col_names = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header)]

    records = []
    for row in rows[1:]:
        date_val = row[2]
        if not isinstance(date_val, datetime):
            continue
        record = {"date": datetime(date_val.year, date_val.month, 1)}
        for c in range(3, len(col_names)):
            slug = "usa_" + col_names[c].lower().replace(" ", "_").replace("(", "").replace(")", "")
            record[slug] = row[c]
        records.append(record)

    return pd.DataFrame(records).sort_values("date").reset_index(drop=True)


def main():
    canada = load_canada()
    uk     = load_uk()
    usa    = load_usa()

    combined = canada.merge(uk, on="date", how="inner").merge(usa, on="date", how="inner")
    combined = combined.sort_values("date").reset_index(drop=True)

    print(f"Overlapping period : {combined['date'].min().strftime('%Y-%m')} to {combined['date'].max().strftime('%Y-%m')} ({len(combined)} months)")
    print(f"Columns            : {list(combined.columns)}")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    combined.to_csv(OUT_PATH, index=False, date_format="%Y-%m-%d")
    print(f"Saved → {OUT_PATH}")


if __name__ == "__main__":
    main()
